"""
BFD Excel Generator
Reproduces the exact layout from the reference BFD format:
  Col B-F  : Inlet streams (IS-xx)
  Col H-J  : Unit Operations block (light blue FFB4C6E7)
  Col K    : separator / arrow
  Col L-O  : Outlet streams (OS-xx) and Process streams (PS-xx)
  Col Q-R  : Temperature / Pressure panel
  Col U-AD : RM Table (right side, rows 2 onward)
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Exact colors from reference file ──────────────────────────────
BLK_BG   = "FFB4C6E7"   # light blue — unit operations block
TITLE_BG = "FFB4C6E7"   # same blue for title row
HDR_BG   = "FFB4C6E7"   # column headers row
YLW_BG   = "FFFFFF00"   # yellow — batch size input cell
GAS_BG   = "FFDAE8FC"   # light blue — gas waste
ORG_BG   = "FFFFF2CC"   # light yellow — organic waste
AQ_BG    = "FFFCE4D6"   # light pink/orange — aqueous waste
SOL_BG   = "FFFFE699"   # light orange — solid waste
PS_BG    = "FFE2EFDA"   # light green — process stream
COMP_BG  = "FFF2F2F2"   # light grey — composition rows
WHITE    = "FFFFFFFF"
NO_FILL  = "00000000"

WASTE_COLOR = {
    "gas_waste"    : GAS_BG,
    "organic_waste": ORG_BG,
    "aqueous_waste": AQ_BG,
    "solid_waste"  : SOL_BG,
}

def _side(style="thin", color="FFAAAAAA"):
    return Side(style=style, color=color)

def _border(top=True, bottom=True, left=True, right=True,
            style="thin", color="FFAAAAAA"):
    s = _side(style, color)
    n = None
    return Border(
        top    = s if top    else n,
        bottom = s if bottom else n,
        left   = s if left   else n,
        right  = s if right  else n,
    )

def _write(ws, row, col, value="", bold=False, bg=None, fg="FF000000",
           halign="left", valign="center", size=11, wrap=True,
           border=None, italic=False, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic, name="Calibri")
    if bg and bg != NO_FILL:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                wrap_text=wrap)
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _merge(ws, r1, c1, r2, c2, value="", bold=False, bg=None,
           fg="FF000000", halign="left", valign="center",
           size=11, wrap=True, border=None, italic=False):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic, name="Calibri")
    if bg and bg != NO_FILL:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                wrap_text=wrap)
    if border:
        cell.border = border
    return cell

# Column index constants (1-based)
cB, cC, cD, cE, cF = 2, 3, 4, 5, 6
cH, cI, cJ         = 8, 9, 10
cK                  = 11
cL, cM, cN, cO     = 12, 13, 14, 15
cQ, cR              = 17, 18
cU, cV, cW, cX, cY, cZ, cAA, cAB, cAC, cAD = 21,22,23,24,25,26,27,28,29,30

def generate_excel(data, output_dir):
    project    = data.get("project", {})
    components = data.get("components", [])
    operations = data.get("operations", [])
    calc       = data.get("calc", {})

    product_name = project.get("product_name", "Product")
    batch_size   = project.get("batch_size", 0)
    reference    = project.get("reference", "")
    date_str     = datetime.now().strftime("%Y-%m-%d")

    calc_comps = calc.get("components", components)
    calc_ops   = calc.get("operations", operations)
    yld        = calc.get("yield", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "BFD"
    ws.sheet_view.showGridLines = False

    # ── Column widths (match reference) ───────────────────────────
    widths = {
        1:3.1, 2:6.5, 3:18.8, 4:39.3, 5:10.5, 6:11.0,
        7:1.5, 8:7.7, 9:33.0, 10:38.1,
        11:10.0, 12:14.4, 13:58.2, 14:10.5, 15:11.0,
        16:1.5, 17:31.7, 18:23.9,
        19:1.5, 20:1.5,
        21:39.1, 22:20.3, 23:28.2, 24:21.9, 25:12.5,
        26:13.5, 27:21.5, 28:20.5, 29:23.0, 30:22.6,
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # All rows default height 17.25
    for r in range(1, 200):
        ws.row_dimensions[r].height = 17.25

    # ── Row 1: Title ───────────────────────────────────────────────
    _merge(ws, 1, cB, 1, cR,
           f"BLOCK FLOW DIAGRAM: {product_name}",
           bold=True, bg=BLK_BG, halign="left", size=11)

    # ── Rows 2-3: Notes / Reference ───────────────────────────────
    _merge(ws, 2, cC, 3, cS if False else 19,
           f"Reference: {reference}   |   Generated: {date_str}",
           bold=False, size=11)

    # ── Rows 4-8: Project summary (left) ──────────────────────────
    _write(ws, 6, cC, "Batch Size", bold=True, size=11)
    _write(ws, 6, cD, batch_size, bold=False, bg=YLW_BG, size=11)
    _write(ws, 6, cE, "kg of SM)", size=11)
    _write(ws, 6, cI, "Output", size=11)
    _write(ws, 6, cJ, yld.get("product_output_kg", ""), size=11)
    _write(ws, 6, cK, "kg of Product", size=11)
    _write(ws, 7, cC, "Yield (Molar %)", bold=False, size=11)
    _write(ws, 7, cD, yld.get("molar_yield_pct", ""), size=11)

    # ── Row 10: Column headers ─────────────────────────────────────
    b = _border(color="FF000000")
    for col, txt in [(cB,""), (cC,"Stream ID"), (cD,"Component"),
                     (cE,"Qty. (kg)"), (cF,"Vol. (kL)")]:
        _write(ws, 10, col, txt, bold=True, bg=BLK_BG, size=11, border=b)
    _merge(ws, 10, cH, 10, cJ, "UNIT OPERATIONS",
           bold=True, bg=BLK_BG, halign="left", size=11, border=b)
    for col, txt in [(cL,"Stream ID"), (cM,"Component"),
                     (cN,"Qty. (kg)"), (cO,"Vol. (kL)")]:
        _write(ws, 10, col, txt, bold=True, bg=BLK_BG, size=11, border=b)
    _write(ws, 10, cQ, "Temperature (°C)", bold=True, bg=BLK_BG, size=11, border=b)
    _write(ws, 10, cR, "Pressure",         bold=True, bg=BLK_BG, size=11, border=b)

    # ── RM Table (right side, starting col U row 2) ────────────────
    rm_row = 2
    _merge(ws, rm_row, cU, rm_row, cAD, "RM Table",
           bold=True, bg=BLK_BG, halign="center", size=11)
    rm_row += 1
    for col, hdr in [(cU,"Component"), (cV,"Mol. wt."), (cW,"MR"),
                     (cX,"No. of mols (kmol)"), (cY,"Purity (%)"),
                     (cZ,"Qty. (kg)"), (cAA,"Density (kg/m3)"),
                     (cAB,"Volume (KL)"), (cAC,"Boiling point(°C)"),
                     (cAD,"Melting Point(°C)")]:
        _write(ws, rm_row, col, hdr, bold=True, bg=BLK_BG, size=10,
               border=_border())
    rm_row += 1
    for c in calc_comps:
        _write(ws, rm_row, cU,  c.get("name",""),          size=10, border=_border())
        _write(ws, rm_row, cV,  c.get("mw",""),            size=10, border=_border(), halign="right")
        _write(ws, rm_row, cW,  c.get("molar_ratio",""),   size=10, border=_border(), halign="right")
        _write(ws, rm_row, cX,  c.get("moles",""),         size=10, border=_border(), halign="right")
        _write(ws, rm_row, cY,  c.get("purity",100),       size=10, border=_border(), halign="right")
        _write(ws, rm_row, cZ,  c.get("mass_kg",""),       size=10, border=_border(), halign="right", bold=True)
        _write(ws, rm_row, cAA, c.get("density",""),       size=10, border=_border(), halign="right")
        _write(ws, rm_row, cAB, c.get("volume_kl",""),     size=10, border=_border(), halign="right")
        _write(ws, rm_row, cAC, c.get("bp",""),            size=10, border=_border())
        _write(ws, rm_row, cAD, c.get("mp",""),            size=10, border=_border())
        rm_row += 1

    # Assumptions block (right side, below RM table)
    rm_row += 1
    _merge(ws, rm_row, cU, rm_row, cAD, "Assumptions",
           bold=True, bg=BLK_BG, halign="left", size=11)
    rm_row += 1
    for op in calc_ops:
        op_type = op.get("type","")
        op_name = op.get("name","")
        if op_type == "reaction":
            _write(ws, rm_row, cU,  f"{op_name} — Conversion (%)", size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("conversion", 100),    size=10, border=_border(), halign="right")
            rm_row += 1
            _write(ws, rm_row, cU,  f"{op_name} — Selectivity (%)", size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("selectivity", 100),   size=10, border=_border(), halign="right")
            rm_row += 1
        elif op_type == "filtration":
            _write(ws, rm_row, cU,  "LOD of Wet Cake (%)",        size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("lod", 30),            size=10, border=_border(), halign="right")
            rm_row += 1
            _write(ws, rm_row, cU,  "Wash Ratio (kg/kg product)", size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("wash_ratio", 2),      size=10, border=_border(), halign="right")
            rm_row += 1
            _write(ws, rm_row, cU,  "Product Loss to ML (%)",     size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("product_loss", 2),    size=10, border=_border(), halign="right")
            rm_row += 1
        elif op_type == "drying":
            _write(ws, rm_row, cU,  "Initial LOD (%)",            size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("lod_initial", 30),    size=10, border=_border(), halign="right")
            rm_row += 1
            _write(ws, rm_row, cU,  "Final LOD (%)",              size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("lod_final", 0.5),     size=10, border=_border(), halign="right")
            rm_row += 1
        elif op_type == "distillation":
            _write(ws, rm_row, cU,  "Distillate Fraction",        size=10, border=_border())
            _write(ws, rm_row, cV,  op.get("distillate_fraction", 0.95), size=10, border=_border(), halign="right")
            rm_row += 1

    # ── BFD Operations (main body, starting row 11) ────────────────
    r = 11
    for op in calc_ops:
        op_type  = op.get("type","")
        op_name  = op.get("name", op_type.title()).upper()
        eq       = op.get("equipment", {})
        steps    = op.get("steps", [])
        inlets   = op.get("inlet_streams", [])
        outlets  = op.get("outlet_streams", [])
        ps       = op.get("process_stream")
        t_init   = op.get("temp_initial","")
        t_final  = op.get("temp_final","")
        pressure = op.get("pressure","atm")

        b = _border(color="FF000000", style="thin")

        block_top = r

        # ── Operation name row ─────────────────────────────────────
        _merge(ws, r, cH, r, cJ, op_name,
               bold=True, bg=BLK_BG, halign="left", size=11, border=b)
        r += 1

        # ── "Process:" label ──────────────────────────────────────
        _merge(ws, r, cH, r, cJ, "Process:",
               bold=True, bg=BLK_BG, halign="left", size=11, border=b)
        r += 1

        # ── Steps ─────────────────────────────────────────────────
        step_start = r
        for i, step in enumerate(steps, 1):
            _write(ws, r, cH, i,    bg=BLK_BG, size=11, border=b)
            _merge(ws, r, cI, r, cJ, step, bg=BLK_BG, size=11,
                   halign="left", border=b)
            r += 1
        if not steps:
            _merge(ws, r, cH, r, cJ, "(no steps defined)",
                   bg=BLK_BG, size=11, border=b)
            r += 1

        # ── Equipment rows ─────────────────────────────────────────
        eq_rows = [
            ("Equipment tag",        eq.get("tag", "—")),
            ("Equipment MOC",        eq.get("moc", "SS-316L")),
            ("Operating Volume, kL", op.get("operating_volume_kl", "—")),
            ("Equipment volume, kL", eq.get("volume_kl", "—")),
        ]
        for label, val in eq_rows:
            _write(ws, r, cH, label, bg=BLK_BG, size=11, border=b)
            _merge(ws, r, cI, r, cJ, val, bg=BLK_BG, size=11,
                   halign="left", border=b)
            r += 1

        # Volume warning (if any)
        if op.get("volume_warning"):
            _merge(ws, r, cH, r, cJ, op["volume_warning"],
                   bg="FFFFF2CC", size=10, halign="left")
            r += 1

        block_bottom = r - 1

        # ── T/P panel (col Q-R), aligned to block top ──────────────
        tp_r = block_top
        _merge(ws, tp_r, cQ, tp_r, cR,
               op_name.title(), bold=True, size=11)
        tp_r += 1
        _write(ws, tp_r, cQ, f"Initial Temp (°C): {t_init}", size=11)
        tp_r += 1
        _write(ws, tp_r, cQ, f"Setpoint Temp (°C): {t_final}", size=11)
        _write(ws, tp_r, cR, pressure, size=11)
        tp_r += 1

        # ── Outlet streams (OS) — col L-O, from tp_r down ─────────
        os_r = tp_r + 1
        for ostr in outlets:
            cat      = ostr.get("category", "organic_waste")
            obg      = WASTE_COLOR.get(cat, ORG_BG)
            ob       = _border(color="FF888888")
            _write(ws, os_r, cL, ostr["id"],            bg=obg, size=10, border=ob, bold=True)
            _write(ws, os_r, cM, ostr["component"],     bg=obg, size=10, border=ob)
            _write(ws, os_r, cN, ostr.get("qty_kg",""), bg=obg, size=10, border=ob, halign="right")
            _write(ws, os_r, cO, ostr.get("vol_kl",""), bg=obg, size=10, border=ob, halign="right")
            os_r += 1
            # Composition sub-rows
            for comp in ostr.get("composition", []):
                _write(ws, os_r, cM, "  " + comp.get("component",""), bg=COMP_BG, size=9, border=ob)
                _write(ws, os_r, cN, comp.get("qty_kg",""),            bg=COMP_BG, size=9, border=ob, halign="right")
                os_r += 1
            if ostr.get("note"):
                _write(ws, os_r, cM, ostr["note"], size=9, italic=True)
                os_r += 1

        # ── Inlet streams (IS) — col B-F, aligned to block_top ────
        is_r = block_top
        for istr in inlets:
            ib = _border(color="FF888888")
            _write(ws, is_r, cC, istr["id"],            size=11, border=ib)
            _write(ws, is_r, cD, istr["component"],     size=11, border=ib)
            _write(ws, is_r, cE, istr.get("qty_kg",""), size=11, border=ib, halign="right")
            _write(ws, is_r, cF, istr.get("vol_kl",""), size=11, border=ib, halign="right")
            # Arrow (col G = 7)
            _write(ws, is_r, 7, "→", bold=True, size=13, halign="center")
            is_r += 1

        # ── Separator + Process stream (PS) — col K-O ─────────────
        _write(ws, r, cK, "-------", size=11, halign="center")
        if ps:
            pb = _border(color="FF2E86C1", style="medium")
            _write(ws, r, cL, ps["id"],          size=11, bold=True, border=pb)
            _write(ws, r, cM, ps["name"],        size=11, bold=True, border=pb)
            _write(ws, r, cN, ps.get("qty_kg",""), size=11, bold=True,
                   border=pb, halign="right")
            _write(ws, r, cO, ps.get("vol_kl",""), size=11, border=pb, halign="right")
            r += 1
            # Composition header
            _write(ws, r, cM, "Tentative Composition", bold=True, bg=PS_BG, size=11,
                   border=_border())
            _write(ws, r, cN, "kg",    bold=True, bg=PS_BG, size=11,
                   border=_border(), halign="right")
            _write(ws, r, cO, "wt%",   bold=True, bg=PS_BG, size=11,
                   border=_border(), halign="right")
            r += 1
            for comp in ps.get("composition", []):
                cb = _border(color="FFAAAAAA")
                _write(ws, r, cM, comp.get("component",""), bg=PS_BG, size=11, border=cb)
                _write(ws, r, cN, comp.get("qty_kg",""),    bg=PS_BG, size=11, border=cb, halign="right")
                _write(ws, r, cO, f"{comp.get('wt_pct','')}%", bg=PS_BG, size=11,
                       border=cb, halign="right")
                r += 1
        else:
            r += 1

        # Gap row between blocks
        r += 2

    # ── Process Yield row ──────────────────────────────────────────
    _write(ws, r, cQ, "Process Yield % wrt SM", bold=True, size=11)
    _write(ws, r, cR, f"{yld.get('molar_yield_pct','')} %", bold=True, size=11,
           bg="FFE2EFDA")

    # ── Sheet 2: Stream Summary ────────────────────────────────────
    ws2 = wb.create_sheet("Stream Summary")
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([12, 30, 14, 12, 20, 25], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    _merge(ws2, 1, 1, 1, 6, "STREAM SUMMARY TABLE",
           bold=True, bg=BLK_BG, halign="center", size=12)
    sr = 3
    for col, hdr in enumerate(["Stream ID","Description","Qty (kg)","Vol (kL)","Category","Notes"], start=1):
        _write(ws2, sr, col, hdr, bold=True, bg=BLK_BG, size=10,
               border=_border(), halign="center")
    sr += 1
    for op in calc_ops:
        for s in op.get("inlet_streams", []):
            _write(ws2, sr, 1, s["id"],                bg="FFD6E4F0", size=10, border=_border(), bold=True)
            _write(ws2, sr, 2, s["component"],         bg="FFD6E4F0", size=10, border=_border())
            _write(ws2, sr, 3, s.get("qty_kg",""),     bg="FFD6E4F0", size=10, border=_border(), halign="right")
            _write(ws2, sr, 4, s.get("vol_kl",""),     bg="FFD6E4F0", size=10, border=_border(), halign="right")
            _write(ws2, sr, 5, "Inlet Stream",         bg="FFD6E4F0", size=10, border=_border())
            sr += 1
        for s in op.get("outlet_streams", []):
            cat = s.get("category","organic_waste")
            bg  = WASTE_COLOR.get(cat, ORG_BG)
            _write(ws2, sr, 1, s["id"],                bg=bg, size=10, border=_border(), bold=True)
            _write(ws2, sr, 2, s["component"],         bg=bg, size=10, border=_border())
            _write(ws2, sr, 3, s.get("qty_kg",""),     bg=bg, size=10, border=_border(), halign="right")
            _write(ws2, sr, 4, s.get("vol_kl",""),     bg=bg, size=10, border=_border(), halign="right")
            _write(ws2, sr, 5, cat.replace("_"," ").title(), bg=bg, size=10, border=_border())
            _write(ws2, sr, 6, s.get("note",""),       bg=bg, size=10, border=_border())
            sr += 1
        ps = op.get("process_stream")
        if ps:
            _write(ws2, sr, 1, ps["id"],               bg="FFE2EFDA", size=10, border=_border(), bold=True)
            _write(ws2, sr, 2, ps["name"],             bg="FFE2EFDA", size=10, border=_border())
            _write(ws2, sr, 3, ps.get("qty_kg",""),    bg="FFE2EFDA", size=10, border=_border(), halign="right")
            _write(ws2, sr, 4, ps.get("vol_kl",""),    bg="FFE2EFDA", size=10, border=_border(), halign="right")
            _write(ws2, sr, 5, "Process Stream",       bg="FFE2EFDA", size=10, border=_border())
            sr += 1
        sr += 1

    # ── Sheet 3: Waste Norms ───────────────────────────────────────
    ws3 = wb.create_sheet("Waste Norms")
    ws3.sheet_view.showGridLines = False
    for i, w in enumerate([25, 20, 22, 30], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    _merge(ws3, 1, 1, 1, 4, "WASTE NORMS",
           bold=True, bg=BLK_BG, halign="center", size=12)
    product_output = yld.get("product_output_kg", 1) or 1

    wr = 3
    for section, cat in [("Organic Waste", "organic_waste"),
                          ("Aqueous Waste", "aqueous_waste"),
                          ("Gas Waste",     "gas_waste"),
                          ("Solid Waste",   "solid_waste")]:
        bg = WASTE_COLOR.get(cat, ORG_BG)
        _merge(ws3, wr, 1, wr, 4, section, bold=True, bg=bg, size=11)
        wr += 1
        for col, hdr in enumerate(["Stream", "Qty (kg)", "Norm (kg/kg Product)", "Remarks"], start=1):
            _write(ws3, wr, col, hdr, bold=True, bg=bg, size=10, border=_border())
        wr += 1
        found = False
        for op in calc_ops:
            for s in op.get("outlet_streams", []):
                if s.get("category","") == cat:
                    norm = round(s.get("qty_kg", 0) / product_output, 4) if product_output else ""
                    _write(ws3, wr, 1, s["component"],      bg=COMP_BG, size=10, border=_border())
                    _write(ws3, wr, 2, s.get("qty_kg",""),  bg=COMP_BG, size=10, border=_border(), halign="right")
                    _write(ws3, wr, 3, norm,                bg=COMP_BG, size=10, border=_border(), halign="right")
                    wr += 1
                    found = True
        if not found:
            _write(ws3, wr, 1, "—", size=10)
            wr += 1
        wr += 1

    # Freeze panes
    ws["B11"].alignment  = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes      = "B11"
    ws2.freeze_panes     = "A4"

    fname = f"BFD_{product_name.replace(' ','_')}_{date_str}.xlsx"
    fpath = os.path.join(str(output_dir), fname)
    wb.save(fpath)
    return fname
